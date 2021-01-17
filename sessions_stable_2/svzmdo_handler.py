# -*- coding: utf-8 -*-

from openpyxl import Workbook, load_workbook
import os
from os import listdir
from os.path import isfile, join
import re, csv
from datetime import datetime
from PyQt5.QtCore import QObject, pyqtSlot, pyqtSignal
import time, math

class Handler(QObject):
    progress = 0
    data = {}
    not_defined = True
    SESSION_REFERENCE = {
        'year': 0,
        'season': ""
    }
    COURSE_LIST = [29, 28, 27, 26, 25, 24, 23, 22, 21, 20]
    COURSE_ORDER = {
        1: None,
        2: None,
        3: None,
        4: None,
        5: None
    }
    SEMESTER_REFERENCE = 1
    start_zmdo = 3
    current_directory = os.path.abspath(os.curdir)
    answer_string = ""
        
    #run_trigger = pyqtSignal()
    progress_trigger = pyqtSignal(int)
    popup_trigger = pyqtSignal(list)
    #popup_trigger = pyqtSignal(str)

    def __init__(self, widget, show_database, is_ready):
        super(Handler, self).__init__() # what's this?
        self.widget = widget

        self.is_ready = is_ready
        #self.show_database = show_database
        #self.run_trigger.connect(self.run) # MARKED LINE
        self.progress_trigger.connect(self.change_progressbar)
        self.popup_trigger.connect(show_database)

        self.ans = None

    def set_answer_string(self, answer_string):
        with open(self.current_directory+"\\database.txt", 'a', encoding='utf-8') as f:
            f.write(answer_string)
                                            

    def change_progressbar(self, progress):
        self.widget.progressBar.setProperty("value", progress)

    @pyqtSlot()
    def run(self):
        self.SV_FOLDERNAME = self.data['foldername_SV']
        self.SESSION_FOLDERNAME = self.data['foldername_SESSION']
        self.firstCourse = self.data['firstCourse']
        self.firstSession = 'сессия-' + self.data['firstSession'] + '.xlsx'
        # getting SV_filenames
        SV_filenames = [f for f in listdir(self.SV_FOLDERNAME) if isfile(join(self.SV_FOLDERNAME, f))]
        # getting SESSION_filenames
        SESSION_filenames = [f for f in listdir(self.SESSION_FOLDERNAME) if isfile(join(self.SESSION_FOLDERNAME, f))]
        
        self.number_of_courses = len(SV_filenames)
        self.sessions_number = len(SESSION_filenames)

        def get_course(SV_FILENAME):
            SV_FILENAME = SV_FILENAME.split('\\')[-1]
            SV_FILENAME = re.sub(r"\D", "", SV_FILENAME)
            return int(SV_FILENAME)

        def get_session_description(SESSION_FILENAME):
            print(SESSION_FILENAME)
            session_description = {
                'year': int(SESSION_FILENAME[12:14]),
                'season': SESSION_FILENAME[7:11]
            }
            return session_description

        def define_references(COURSE, SESSION_FILENAME):
            # DEFINIG COURSE ORDER
            for i in range(10):
                if self.COURSE_LIST[i] == self.firstCourse:
                    ind = i + 1
                    break
            self.COURSE_ORDER[1] = self.firstCourse
            
            for num in range(2, 6):
                if ind < 10:
                    self.COURSE_ORDER[num] = self.COURSE_LIST[ind]
                    ind += 1
                else:
                    ind = 0
                    self.COURSE_ORDER[num] = self.COURSE_LIST[ind]
                    ind += 1
            #self.SESSION_REFERENCE = get_session_description(SESSION_FILENAME)
            self.SESSION_REFERENCE = get_session_description(self.firstSession)

        def get_dict_difference(course):
            for key, value in self.COURSE_ORDER.items():
                if value == course:
                    return key-1 # -1 cause we're looking for a difference according to the 1st course

        def get_semester(SESSION_FILENAME, COURSE):
            # according to SESSION
            session_description = get_session_description(SESSION_FILENAME)
            #print(session_description)
            semester = self.SEMESTER_REFERENCE
            semester += 2*(session_description['year'] - self.SESSION_REFERENCE['year'])
            if self.SESSION_REFERENCE['season'] == 'зима':
                if session_description['season'] == 'лето':
                    semester += 1
            else:
                if session_description['season'] == 'зима':
                    semester -= 1

            # according to COURSE
            difference_in_courses = get_dict_difference(COURSE) # IT CALCULATES THE DIFFERENCE ACCORDING TO THE 1ST COURSE, SO WE GET KINDA A BIAS IN SEMESTERS
            semester += 2*difference_in_courses
            # a piece of godlike debug code here
            #print(SESSION_FILENAME)
            #print(semester)
            #print('diff_in_courses', difference_in_courses)
            #print()
            return semester

        #progress
        self.progress_sv = int(70 / self.number_of_courses)
        self.sv_number = len(SV_filenames)

        # opening SV file
        for SV_FILENAME in SV_filenames:
        
            sdm_list = {1:{},2:{},3:{},4:{},5:{},6:{},7:{},8:{},9:{},10:{}}
            # 1 list - 1 course from SV
            # 1 course - 10 semesters
            # 1 semester - 1 list                       with X elements, where X is the number of st.groups in the course
            # 1 such st.group list - X sdm_dict objects
            # 1 sdm_dict object - Y cadets
            # 1 cadet - Z disciplines, including avg mark for the session
            # 1 discipline - 1 mark

            # [semester][NAME of study froum SHEET.TITLE][surname][discipline] -> mark

            column_average_marks = {}

            remove_lower = lambda text: re.sub('[а-я]', '', text)
            def get_surname_initials(name):
                index_of_second_upper = 1
                for ind in range(2, len(name)):
                    if name[ind].isupper():
                        index_of_second_upper = ind
                        break
                second_part_of_name = remove_lower(name[index_of_second_upper+1:])
                name = name[0:index_of_second_upper+1] + second_part_of_name
                    
                name = name.replace(' ', '')
                name = name.replace('.', '')
                name = name.replace(',', '')
                name = name.replace('\n', '')

                return name

            # loading SV file
            wb_SV = load_workbook(join(self.SV_FOLDERNAME, SV_FILENAME), data_only=True)
            # getting COURSE from SV filename
            COURSE = get_course(SV_FILENAME)

            # and creating of name for it and then saving it
            now = datetime.now()
            output_filename = now.strftime("%d.%m.%Y_%H-%M-%S.xlsx")
            output_filename = self.current_directory + '\\documents\\Вывод\\' + 'Сводная ведомость (' + str(COURSE) + ' курс) ' + output_filename
            wb_SV.save(output_filename)

            SV_sheetnames = wb_SV.sheetnames
            # getting the order of courses. It's counted once
            if self.not_defined:
                define_references(COURSE, SESSION_filenames[0])
                # self.SEMESTER_REFERENCE = -3 # it seems to me that it should be 1 cause we're looking at the 1st course every time
                self.not_defined = False

            # opening SESSION file
            for SESSION_filename in SESSION_filenames:
                # loading file
                wb_SESSION = load_workbook(join(self.SESSION_FOLDERNAME, SESSION_filename), data_only=True)
                # get the SEMESTER number
                SEMESTER = get_semester(SESSION_filename, COURSE)

                if 1 <= SEMESTER and SEMESTER <= 10:
                    # for study groups...
                    for sheet in wb_SESSION:
                        # if there's a conformity between SV and SESSIONS sheets...
                        if str(COURSE) in sheet.title and sheet.title in SV_sheetnames:
                            
                            # FILLING ITSELF
                            ws_group_session = sheet
                            ws_group_sv = wb_SV[sheet.title]

                            # getting list of disciplines from SESSIONS
                            disciplines = []
                            shift = []
                            r = 5
                            c = 6
                            breakpoint = 0
                            
                            while True:
                                val = ws_group_session.cell(row=r, column=c).value
                                if type(val) != str:
                                    breakpoint += 1
                                    if breakpoint > 15:
                                        break
                                else:
                                    disciplines.append(str(val).replace(' ', '').lower())
                                    shift.append(1+breakpoint)
                                    breakpoint = 0
                                c += 1
                                
                            # SDM is 'surname - discipline - mark'
                            sdm_dict = {}
                            # getting list of surnames
                            c = 4
                            # FILLING OF SDM WITH MARKS ITSELF
                            for r in range(6, 40):
                                # val is SURNAME
                                val = ws_group_session.cell(row=r, column=c).value
                                if type(val) == str:
                                    if len(val) > 3:
                                        #val = val.split()[0]
                                        val = get_surname_initials(val)
                                        sdm_dict[val] = dict()
                                        # appending marks to SDM dict
                                        c_m = 6
                                        num = 0
                                        for discipline in disciplines:
                                            mark = ws_group_session.cell(row=r, column=c_m).value
                                            # val is SURNAME
                                            if c_m == 30: # 'AD' column is under number 30
                                                sdm_dict[val]["average_mark"] = mark
                                            else:
                                                sdm_dict[val][discipline] = mark
                                            
                                            c_m += shift[num]
                                            num += 1
                                        
                            sdm_list[SEMESTER][sheet.title] = sdm_dict
                            #print('Ive finished sdm_list[', SEMESTER,'][',sheet.title,']')
                            #print(sdm_dict)
                            #for keys,values in sdm_dict.items():
                            #    print('SEMESTER', SEMESTER)
                            #    print(keys)
                            #    print(values)
                            r = 4
                            c = 3
                            # Here we're trying to catch the number of column containing AVG_MARKS:
                            column_average_marks[sheet.title] = 0
                            while True:
                                val = ws_group_sv.cell(row=r, column=c).value
                                if type(val) != str and type(val) != int:
                                    column_average_marks[sheet.title] = c
                                    break
                                c += 1

                            # filling the SV
                            r = 4
                            c = 3
                            breakpoint = 0
                            
                            # now let's scroll through the line in SV containing numbers of semesters...
                            while True:
                                val = ws_group_sv.cell(row=r, column=c).value
                                if type(val) != str and type(val) != int:
                                    breakpoint += 1
                                    if breakpoint > 10:
                                        break
                                else:
                                    breakpoint = 0
                                    # if semester in SV is the same to the considering SESSION file:
                                    if int(val) == SEMESTER:
                                        # check if this discipline from SESSIONS is in the database
                                        def in_database_check(discipline_SV):
                                            with open(self.current_directory+"\\database.txt", 'r', encoding='utf-8') as f:
                                                lines = f.readlines()
                                                for line in lines:
                                                    if discipline_SV in line:
                                                        return_string = line.replace(discipline_SV, '')
                                                        return_string = return_string.replace(';', '')
                                                        return_string = return_string.replace('\n', '')
                                                        #print('WOW! RETURN_STRING:', return_string)
                                                        #print()
                                                        if return_string in disciplines:
                                                            return (True, return_string)
                                            return (False, None)

                                        def mark_writer(discipline):
                                            #scroll through cadets
                                            for r_m in range (6, 40):
                                                surname = str(ws_group_sv.cell(row=r_m, column=2).value)
                                                if len(surname) > 3:
                                                    #surname = surname.split()[0]
                                                    surname = get_surname_initials(surname)
                                                    if surname in sdm_dict:
                                                        #print('Ive tried', discipline)
                                                        #print()
                                                        ws_group_sv.cell(row=r_m, column=c).value = sdm_dict[surname][discipline.replace(' ', '').lower()]
                                                        #print('Ive changed it!')

                                        # get the discipline name from SV
                                        discipline_SV = str(ws_group_sv.cell(row=r+1, column=c).value).replace(' ', '').lower()
                                        # check for the conformity
                                        if discipline_SV in disciplines:
                                            mark_writer(discipline_SV)
                                        else:
                                            while True:
                                                db = in_database_check(discipline_SV)
                                                # if it's already in the database:
                                                if db[0]:
                                                    mark_writer(db[1])
                                                    break
                                                else:
                                                    list_with_disciplines = [discipline_SV, disciplines]
                                                    self.popup_trigger.emit(list_with_disciplines)
                                                    #self.popup_trigger.emit("lol")
                                                    self.is_ready.wait()
                                                    self.is_ready.clear()
                                                    
                                                    print('IT\'S TIME FOR DATABASE POP-UP!')
                                                    print("Semester number:", SEMESTER)
                                                    print("Discipline_SV name:", discipline_SV)
                                                    print('SESSION filename:', SESSION_filename)
                                                    print('GROUP number:', sheet.title)
                                                    print()
                                                    # ЗАБИВАТЬ В БД ИЗ comboBox ТОЛЬКО LOWERCASE!!! 
                                        
                                c += 1
                #progress
                self.progress += int(self.progress_sv / self.sessions_number)
                self.progress_trigger.emit(self.progress)

            # LETS COUNT AVG FOR THIS SV OF SOME STUDY GROUP:
            SV_sheetnames = wb_SV.sheetnames
            for SV_sheetname in SV_sheetnames:
                ws_group_sv = wb_SV[SV_sheetname]
                for r in range(6, 40):
                    count = 0
                    summ = 0
                    if SV_sheetname in column_average_marks.keys():
                        for c in range(3, column_average_marks[SV_sheetname]):
                            mark = str(ws_group_sv.cell(row=r, column=c).value)
                            mark = mark[-1]
                            #print(mark)
                            if mark.isnumeric():
                                mark = int(mark)
                                summ += mark
                                count += 1
                    if count != 0:
                        average_grade_point = summ / count
                        if SV_sheetname in column_average_marks.keys():
                            ws_group_sv.cell(row=r, column=column_average_marks[SV_sheetname]).value = str(average_grade_point)[0:5]

            wb_SV.save(output_filename)

            #progress After this - less than 25 percents left
            self.progress += 5
            self.progress_trigger.emit(self.progress)


            # HERE ZMDO STARTS

            # loading ZMDO template
            wb_zmdo = load_workbook(self.current_directory + "\\documents\\шаблон ЗМ ДО 2 факультет.xlsx")
            # creating an Excel workspace
            ws_zmdo = wb_zmdo.active
            #ws_zmdo = wb_zmdo.copy_worksheet(ws_zmdo)
            ws_zmdo.title = str(COURSE) + " курс"

            now = datetime.now()
            output_filename_zmdo = now.strftime("%d.%m.%Y_%H-%M-%S.xlsx")
            output_filename_zmdo = self.current_directory + '\\documents\\Вывод\\'+'Данные ЗМ ДО (' + str(COURSE) + ' курс) ' + output_filename_zmdo
            wb_zmdo.save(output_filename_zmdo)

            SV_sheetnames = wb_SV.sheetnames
            # for every study froup sheet in SV:
            for study_group_sheetname in SV_sheetnames:
                ws_study_group = wb_SV[study_group_sheetname]

                # name == 'ФамилияИО'
                excellent_student = {}
                bad = 0
                
                # fill NAME, AVG and STRAIGHT A parameters for our excellent students
                for r in range (6, 40):
                    if study_group_sheetname in column_average_marks.keys():
                        avg = ws_study_group.cell(row=r, column=column_average_marks[study_group_sheetname]).value
                        if type(avg) != type(None):
                            #if type(avg) == type(str) and avg.replace(',','').replace('.','').isnumeric():
                            avg = float(avg)
                            if avg >= 4.6:
                                name = str(ws_study_group.cell(row=r, column=2).value)
                                name = get_surname_initials(name)

                                excellent_student[name] = {
                                    'average_mark': [0.0] * 11,
                                    'is_straight_A': False,
                                    'note': '',
                                    '2': False,
                                    '3': 0,
                                    '4': 0,
                                    '5': 0
                                }

                                excellent_student[name]['average_mark'][0] = avg
                                if avg >= 4.9:
                                    excellent_student[name]['is_straight_A'] = True
                
                for name_excellent in excellent_student.keys():
                    for semester_number in sdm_list.keys():
                        #print(sdm_list[semester_number].keys()) ###
                        if study_group_sheetname in sdm_list[semester_number].keys():
                            group_list_in_sdm = sdm_list[semester_number][study_group_sheetname]
                            for name_in_sdm in group_list_in_sdm.keys():
                                if name_excellent == name_in_sdm:
                                    # What the hell is below?
                                    #excellent_student[name_excellent] = excellent_student[name_excellent]
                                    #group_list_in_sdm[name_in_sdm] = group_list_in_sdm[name_in_sdm]
                                    excellent_student[name_excellent]['average_mark'][semester_number] = group_list_in_sdm[name_in_sdm]['average_mark']

                                    goal_marks = [3]
                                    if excellent_student[name_excellent]['is_straight_A']:
                                        goal_marks.append(4)

                                    for discipline_name_sdm in group_list_in_sdm[name_in_sdm].keys():
                                        mark = group_list_in_sdm[name_in_sdm][discipline_name_sdm]
                                        if type(mark) == str:
                                            mark = mark.replace(' ', '')[-1]
                                            print(mark, '\n')
                                            if mark.isnumeric():
                                                mark = int(mark[-1])
                                        if type(mark) == int:
                                            if not excellent_student[name_excellent]['2']:
                                                if mark == 2:   
                                                    excellent_student[name_excellent]['2'] = True
                                                    bad += 1
                                                elif mark in goal_marks:
                                                    if mark == 3:
                                                        excellent_student[name_excellent]['3'] += 1
                                                    else:
                                                        excellent_student[name_excellent]['4'] += 1
                                                    old_note = excellent_student[name_excellent]['note']
                                                    new_note = old_note + '"' + str(mark) + '" - ' + discipline_name_sdm + " (" + str(semester_number) + " сем.)\n"
                                                    excellent_student[name_excellent]['note'] = new_note
                                                else:
                                                    excellent_student[name_excellent]['5'] += 1
                        else:
                            print('STUDY GROUP', study_group_sheetname,'isnt found in SDM')
                            print('SEMESTER:', semester_number)
                            print('KEYS:', sdm_list[semester_number].keys())
                            print()
                            #pass

                # LOADING THE INFO TO EXCEL
                additional = len(excellent_student.keys()) - bad
                ws_zmdo.insert_rows(self.start_zmdo+1, amount=additional)

                r = self.start_zmdo
                ws_zmdo.cell(row=r, column=3).value = "Учебная группа " + study_group_sheetname
                #print("Ive started filling group N", study_group_sheetname)
                    
                for student_name in excellent_student.keys():
                    student = excellent_student[student_name]
                    if not student['2'] and student['3'] <= 3 and student['4'] <= 16 and student_name != 'None':
                        r += 1
                        ws_zmdo.cell(row=r, column=3).value = student_name

                        avg_of_avgs = 0
                        avg_count = 0
                        #avg for sessions
                        for i in range(1, 11):
                            avg_to_cell = str(student['average_mark'][i])
                            if avg_to_cell.replace('.', '').isnumeric():
                                if float(avg_to_cell) == 0.0:
                                    continue
                                avg_to_cell = str(student['average_mark'][i])[0:5]
                                ws_zmdo.cell(row=r, column=3+i).value = avg_to_cell
                                avg_of_avgs += float(avg_to_cell)
                                avg_count += 1

                        #avg for avg for sessions
                        avg_of_avgs = avg_of_avgs / avg_count
                        avg_of_avgs = str(avg_of_avgs)[0:5]
                        ws_zmdo.cell(row=r, column=14).value = avg_of_avgs

                        #avg for diploma
                        ws_zmdo.cell(row=r, column=15).value = student['average_mark'][0]

                        #numbers of marks
                        ws_zmdo.cell(row=r, column=16).value = student['5']
                        ws_zmdo.cell(row=r, column=17).value = student['4']
                        ws_zmdo.cell(row=r, column=18).value = student['3']
                        ws_zmdo.cell(row=r, column=19).value = 0

                        #note
                        ws_zmdo.cell(row=r, column=21).value = student['note']
                
                self.start_zmdo = r+1
                wb_zmdo.save(output_filename_zmdo)
                
                for keys,values in excellent_student.items():
                    print("sdm_list[", semester_number, "].keys() = ", sdm_list[semester_number].keys())
                    print('GROUP', study_group_sheetname)
                    print(keys)
                    print(values)
                    print()
                
                # pr
                self.progress += int(24 / self.sv_number)
                self.progress_trigger.emit(self.progress)

            self.start_zmdo = 3


        self.progress = 100
        self.progress_trigger.emit(self.progress)
        print('ITS ALL DONE')
